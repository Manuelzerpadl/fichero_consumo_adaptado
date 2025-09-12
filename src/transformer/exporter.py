from openpyxl import load_workbook
from pathlib import Path
import pandas as pd


def export_with_config(df: pd.DataFrame, config: dict, output_file: Path) -> None:
    """
    Exporta un DataFrame a Excel con reglas específicas en output_config.
    """
    output_cfg = config["output_config"]

    # Guardar DataFrame en Excel (pandas -> openpyxl)
    writer = pd.ExcelWriter(output_file, engine="openpyxl")
    df.to_excel(
        writer,
        index=False,
        sheet_name=output_cfg.get("EXCEL - Nombre hoja", "Sheet1"),
        startrow=int(output_cfg.get("EXCEL - Fila inicio datos", "7")) - 1,
    )
    writer.close()

    # Reabrir con openpyxl para meter valores en celdas específicas
    wb = load_workbook(output_file)
    ws = wb[output_cfg.get("EXCEL - Nombre hoja", "Sheet1")]

    if "EXCEL - Celda C2" in output_cfg:
        ws["C2"] = output_cfg["EXCEL - Celda C2"]
    if "EXCEL - Celda J2" in output_cfg:
        ws["J2"] = output_cfg["EXCEL - Celda J2"]
    if "EXCEL - Celda B7" in output_cfg:
        ws["B7"] = output_cfg["EXCEL - Celda B7"]

    wb.save(output_file)
